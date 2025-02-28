from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
import requests
from datetime import datetime
import pandas as pd
import xml.etree.ElementTree as ET
import html
import os
import glob
from telegram_utils import send_to_telegram
import openpyxl

TELEGRAM_TOKEN = "7810323512:AAEL6hDjjZgz64gADrJfcKwrqO42himl3oI"
TELEGRAM_CHAT_ID = "-4675102066"  # Chuyển sang string để đảm bảo định dạng đúng

def process_connection_points(excel_file):
    print("Đang xử lý thông tin điểm đầu - điểm cuối...")
    # Đọc file Excel
    df = pd.read_excel(excel_file)
    
    # Tạo list để lưu dữ liệu điểm đầu - điểm cuối
    connection_data = []
    
    # Định nghĩa các từ khóa cần lọc
    keywords = ['BVI', 'PTO', 'STY', 'TTT', 'DPG']
    
    # Xử lý từng dòng trong cột sensor_raw
    for sensor in df['sensor_raw']:
        if isinstance(sensor, str) and 'To' in sensor:
            # Tách chuỗi tại 'To'
            parts = sensor.split('To')
            if len(parts) >= 2:
                diem_dau_full = parts[0].strip()
                diem_cuoi = parts[1].strip()
                
                # Tách loại uplink và điểm đầu
                diem_dau_parts = diem_dau_full.split(' ', 1)
                loai_uplink = diem_dau_parts[0] if len(diem_dau_parts) > 1 else ''
                diem_dau = diem_dau_parts[1] if len(diem_dau_parts) > 1 else diem_dau_full
                
                # Loại bỏ dấu '-' và nội dung sau nó ở điểm cuối nếu có
                if ' - ' in diem_cuoi:
                    diem_cuoi = diem_cuoi.split(' - ')[0].strip()
                
                # Kiểm tra xem điểm đầu hoặc điểm cuối có chứa một trong các từ khóa không
                if any(keyword in diem_dau or keyword in diem_cuoi for keyword in keywords):
                    connection_data.append({
                        'Loại uplink': loai_uplink,
                        'Điểm đầu': diem_dau,
                        'Điểm cuối': diem_cuoi
                    })
    
    # Tạo DataFrame mới cho điểm đầu - điểm cuối
    df_connections = pd.DataFrame(connection_data)
    
    # Tạo ExcelWriter object để thêm sheet mới
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
        # Thêm sheet mới
        df_connections.to_excel(writer, sheet_name='Điểm đầu - Điểm cuối', index=False)
    
    print("Đã tạo sheet mới 'Điểm đầu - Điểm cuối' trong file Excel!")

def xml_to_excel(xml_file):
    print("Đang xử lý file XML...")
    # Đọc file XML
    tree = ET.parse(xml_file)
    root = tree.getroot()
    
    # Khởi tạo list để lưu dữ liệu
    data = []
    
    # Duyệt qua các phần tử item trong XML
    for item in root.findall('.//item'):
        # Lấy dữ liệu từ các thẻ con và xử lý HTML entities
        row_data = {
            'sensor_raw': html.unescape(item.find('sensor_raw').text) if item.find('sensor_raw') is not None else '',
            'status': html.unescape(item.find('status').text) if item.find('status') is not None else '',
            'status_raw': item.find('status_raw').text if item.find('status_raw') is not None else '',
            'lastvalue': item.find('lastvalue').text if item.find('lastvalue') is not None else '',
            'priority': item.find('priority').text if item.find('priority') is not None else '',
            'objid': item.find('objid').text if item.find('objid') is not None else '',
            'baselink_raw': item.find('baselink_raw').text if item.find('baselink_raw') is not None else ''
        }
        data.append(row_data)
    
    # Tạo DataFrame từ dữ liệu
    df = pd.DataFrame(data)
    
    # Tạo tên file Excel với timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file = f"data_{timestamp}.xlsx"
    
    # Lưu DataFrame vào file Excel
    df.to_excel(excel_file, index=False, sheet_name='Raw Data')
    print(f"Đã lưu dữ liệu vào file Excel: {excel_file}")
    
    # Xử lý và tạo sheet điểm đầu - điểm cuối
    process_connection_points(excel_file)
    
    return excel_file

def cleanup_old_files():
    print("Đang dọn dẹp các file cũ...")
    
    # Xóa tất cả file XML cũ trừ file mới nhất
    xml_files = glob.glob("data_*.xml")
    if xml_files:
        # Sắp xếp theo thời gian tạo file, mới nhất ở cuối
        xml_files.sort(key=os.path.getctime)
        # Xóa tất cả file trừ file cuối cùng (mới nhất)
        for file in xml_files[:-1]:
            try:
                os.remove(file)
                print(f"Đã xóa file: {file}")
            except Exception as e:
                print(f"Không thể xóa file {file}: {str(e)}")
    
    # Xóa tất cả file Excel cũ trừ file mới nhất
    excel_files = glob.glob("data_*.xlsx")
    if excel_files:
        # Sắp xếp theo thời gian tạo file, mới nhất ở cuối
        excel_files.sort(key=os.path.getctime)
        # Xóa tất cả file trừ file cuối cùng (mới nhất)
        for file in excel_files[:-1]:
            try:
                os.remove(file)
                print(f"Đã xóa file: {file}")
            except Exception as e:
                print(f"Không thể xóa file {file}: {str(e)}")
    
    print("Đã hoàn tất dọn dẹp file!")

def process_excel_data(file_path):
    try:
        print("Bắt đầu xử lý file Excel...")
        # Đọc file Excel
        wb = openpyxl.load_workbook(file_path)
        raw_sheet = wb['Raw Data']
        
        # Danh sách keywords và tạo sheets mới
        keywords = ['BVI', 'PTO', 'STY', 'TTT', 'DPG']
        
        # Xóa các sheet cũ nếu tồn tại
        for keyword in keywords:
            if keyword in wb.sheetnames:
                wb.remove(wb[keyword])
            
            # Tạo sheet mới
            new_sheet = wb.create_sheet(title=keyword)
            print(f"Đã tạo sheet mới: {keyword}")
            
            # Tạo tiêu đề cho các sheet
            new_sheet['A1'] = 'Điểm đầu'
            new_sheet['B1'] = 'Điểm cuối'
            if keyword == 'STY':
                new_sheet['C1'] = 'Loại uplink'
        
        # Xử lý dữ liệu từ Raw Data
        row_count = {k: 2 for k in keywords}  # Bắt đầu từ hàng 2 cho mỗi sheet
        
        sensor_raw_column = None
        # Tìm cột sensor_raw
        for cell in raw_sheet[1]:
            if cell.value == 'sensor_raw':
                sensor_raw_column = cell.column
                break
                
        if not sensor_raw_column:
            raise Exception("Không tìm thấy cột sensor_raw trong sheet Raw Data")
            
        print("Bắt đầu xử lý dữ liệu từ sensor_raw...")
        
        for row in raw_sheet.iter_rows(min_row=2):
            sensor_raw = row[sensor_raw_column - 1].value
            if not sensor_raw:
                continue
                
            for keyword in keywords:
                if keyword in str(sensor_raw):
                    sheet = wb[keyword]
                    current_row = row_count[keyword]
                    
                    # Xử lý đặc biệt cho sheet STY
                    if keyword == 'STY':
                        # Tìm điểm đầu (chứa STY)
                        start_point = ''
                        for word in str(sensor_raw).split():
                            if 'STY' in word:
                                start_point = word
                                break
                        
                        # Tìm điểm cuối (sau từ "To")
                        end_point = ''
                        if 'To' in sensor_raw:
                            parts = sensor_raw.split('To')
                            if len(parts) > 1:
                                end_point = parts[1].strip().split()[0]
                        
                        # Xác định loại uplink
                        uplink_type = ''
                        if '1G' in sensor_raw:
                            uplink_type = '1G'
                        elif '10G' in sensor_raw:
                            uplink_type = '10G'
                        
                        sheet[f'A{current_row}'] = start_point
                        sheet[f'B{current_row}'] = end_point
                        sheet[f'C{current_row}'] = uplink_type
                    
                    # Xử lý cho các sheet khác
                    else:
                        # Tìm điểm đầu (chứa keyword)
                        start_point = ''
                        for word in str(sensor_raw).split():
                            if keyword in word:
                                start_point = word
                                break
                        
                        # Tìm điểm cuối (sau từ "To")
                        end_point = ''
                        if 'To' in sensor_raw:
                            parts = sensor_raw.split('To')
                            if len(parts) > 1:
                                end_point = parts[1].strip().split()[0]
                        
                        sheet[f'A{current_row}'] = start_point
                        sheet[f'B{current_row}'] = end_point
                    
                    row_count[keyword] += 1
        
        print("Đang lưu file Excel...")
        wb.save(file_path)
        print("Đã xử lý xong file Excel")
        return True
    except Exception as e:
        print(f"Lỗi khi xử lý Excel: {str(e)}")
        return False

def process_and_send_data(driver, cookies_dict):
    try:
        # URL của file XML
        xml_url = "http://203.210.142.155/api/table.xml?&content=sensors&columns=probegroupdevice%3Dtext%2Csensor%3Dhtmllong%2Cstatus%2Cmessage%2Clastvalue%2Cminigraph%2Cpriority%2Cfavorite%2Ccheckbox%2Cobjid%2Cbaselink&sortby=priority&tabletitle=AUTO&filter_status=5&id=0"
        
        print("Đang tải file XML...")
        # Tải file XML sử dụng requests với cookies từ phiên đăng nhập
        response = requests.get(xml_url, cookies=cookies_dict)
        
        # Tạo tên file với timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        xml_filename = f"data_{timestamp}.xml"
        
        # Lưu nội dung XML vào file
        with open(xml_filename, 'wb') as f:
            f.write(response.content)
        print(f"Đã tải file XML thành công! Lưu tại: {xml_filename}")
        
        # Chuyển đổi XML sang Excel
        excel_file = xml_to_excel(xml_filename)
        print(f"Đã chuyển đổi XML sang Excel thành công!")
        
        # In đường dẫn đầy đủ của file
        abs_excel_file = os.path.abspath(excel_file)
        print(f"Đường dẫn đầy đủ của file Excel: {abs_excel_file}")
        
        # Sau khi download file, thêm xử lý Excel
        if os.path.exists(abs_excel_file):
            process_excel_data(abs_excel_file)
        
        # Gửi file Excel đến Telegram
        if send_to_telegram(abs_excel_file, TELEGRAM_TOKEN, TELEGRAM_CHAT_ID):
            print("Gửi file thành công!")
        else:
            print("Không thể gửi file!")
        
        # Dọn dẹp các file cũ
        cleanup_old_files()
        
        print(f"Hoàn tất xử lý lúc: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        return True
        
    except Exception as e:
        print(f"Lỗi trong quá trình xử lý và gửi dữ liệu: {str(e)}")
        return False

try:
    print("Bắt đầu khởi tạo trình duyệt...")
    # Thay đổi cách khởi tạo Service
    service = Service(ChromeDriverManager().install())

    # Tạo options cho Chrome
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.binary_location = os.getenv('CHROME_BIN', '/usr/bin/chromium')

    # Khởi tạo trình duyệt Chrome với Service và Options đã cấu hình
    driver = webdriver.Chrome(service=service, options=chrome_options)
    print("Đã khởi tạo trình duyệt thành công!")

    print("Đang truy cập trang web...")
    # Mở trang web
    driver.get("http://203.210.142.155/")
    print("Đã mở trang web thành công!")
    
    print("Đang đợi trường username xuất hiện...")
    # Đợi cho đến khi trường username xuất hiện
    username_field = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "loginusername"))
    )
    print("Đã tìm thấy trường username!")
    
    print("Đang tìm các phần tử đăng nhập...")
    # Tìm trường password
    password_field = driver.find_element(By.ID, "loginpassword")
    
    # Tìm nút đăng nhập
    login_button = driver.find_element(By.ID, "submitter1")
    print("Đã tìm thấy tất cả các phần tử đăng nhập!")
    
    print("Đang điền thông tin đăng nhập...")
    # Điền thông tin đăng nhập
    username_field.send_keys("omc")
    password_field.send_keys("omc2014")
    print("Đã điền thông tin đăng nhập!")
    

    print("Đang click nút đăng nhập...")
    # Click nút đăng nhập
    login_button.click()
    print("Đã click nút đăng nhập!")
    
    print("Đang chờ đăng nhập hoàn tất...")
    time.sleep(2)  # Đợi 2 giây để đảm bảo đăng nhập hoàn tất
    
    # Lấy cookies từ trình duyệt
    cookies = driver.get_cookies()
    cookies_dict = {cookie['name']: cookie['value'] for cookie in cookies}
    
    print("Bắt đầu vòng lặp xử lý định kỳ...")
    while True:
        print(f"\nBắt đầu xử lý lúc: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        if process_and_send_data(driver, cookies_dict):
            print("Đang chờ 5 phút cho lần xử lý tiếp theo...")
            time.sleep(300)  # Đợi 5 phút
        else:
            print("Có lỗi xảy ra, thử lại sau 1 phút...")
            time.sleep(60)  # Đợi 1 phút nếu có lỗi
    
except Exception as e:
    print(f"Có lỗi xảy ra: {str(e)}")
    print("Chi tiết lỗi:", e.__class__.__name__)
    input("Nhấn Enter để đóng trình duyệt...")
    driver.quit()